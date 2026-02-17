from __future__ import annotations

from pathlib import Path
import sys
import shutil
import configparser
import pandas as pd
from openpyxl import load_workbook

from functions import edgar_workflow, reporting


def _resolve_path(project_root: Path, maybe_path: str) -> Path:
    """Resolve absolute path from a possibly relative path in config.ini."""
    p = Path(maybe_path)
    if not p.is_absolute():
        p = (project_root / p).resolve()
    return p


def _ric_to_edgar(ric: str) -> str:
    """Convert a Refinitiv RIC (e.g., 'AAPL.OQ', 'BRKb.N') to an EDGAR-friendly ticker.

    Rules:
    - If a slash is present (e.g., 'BF/B'), convert to hyphen form ('BF-B').
    - If a dot is present, keep the prefix before the dot.
    - If the prefix ends with a lowercase letter (e.g., 'BRKb'), convert to 'BRK-B'.
    - Uppercase the result.
    """
    s = str(ric).strip()
    if not s:
        return ""
    # Handle share-class tickers with slash (BF/B, BRK/B) → BF-B, BRK-B
    if "/" in s:
        parts = s.split("/", 1)
        return f"{parts[0].upper()}-{parts[1].upper()}"
    if "." in s:
        pre, _ = s.split(".", 1)
    else:
        pre = s
    if pre and pre[-1].islower():
        pre = f"{pre[:-1]}-{pre[-1].upper()}"
    return pre.upper()


def main() -> None:
    project_root = Path(__file__).resolve().parent
    config_path = project_root / "config.ini"
    if not config_path.exists():
        print(f"[ERROR] config.ini not found at {config_path}")
        sys.exit(1)

    config = configparser.ConfigParser()
    try:
        config.read(config_path, encoding="utf-8")
    except Exception as e:
        print(f"[ERROR] Failed to read config.ini: {e}")
        sys.exit(1)

    # Load companies CSV path
    try:
        companies_csv_rel = config["General"]["companies_csv_path"]
    except KeyError:
        print("[ERROR] Missing 'General.companies_csv_path' in config.ini")
        sys.exit(1)

    companies_csv_path = _resolve_path(project_root, companies_csv_rel)
    if not companies_csv_path.exists():
        print(f"[ERROR] companies.csv not found at {companies_csv_path}")
        sys.exit(1)

    # Load companies
    try:
        df = pd.read_csv(companies_csv_path)
    except Exception as e:
        print(f"[ERROR] Failed to read companies CSV: {e}")
        sys.exit(1)

    if "Ticker" not in df.columns:
        print("[ERROR] companies.csv must contain a 'Ticker' column")
        sys.exit(1)

    # Accept RICs (e.g., AAPL.O, MSFT.OQ) and convert to EDGAR tickers for downloads
    df["RIC"] = df["Ticker"].astype(str).str.strip()
    df["EDGAR_Ticker"] = df["RIC"].apply(_ric_to_edgar)

    # Build mapping EDGAR_Ticker -> RIC (first occurrence wins) and a unique EDGAR list
    edgar_to_ric = {}
    edgar_list = []
    for edg, ric in zip(df["EDGAR_Ticker"], df["RIC"]):
        edg = str(edg).strip().upper()
        ric = str(ric).strip().upper()
        if not edg:
            continue
        if edg not in edgar_to_ric:
            edgar_to_ric[edg] = ric
        if edg not in edgar_list:
            edgar_list.append(edg)

    print(f"[OK] Loaded {len(edgar_list)} companies from {companies_csv_path.name}")
    if edgar_list:
        preview = ", ".join([f"{edgar_to_ric[e]}→{e}" for e in edgar_list[:10]])
        print(f"[Preview] {preview}{' ...' if len(edgar_list) > 10 else ''}")

    # Phase 6: Run workflow and generate report
    if not edgar_list:
        print("[WARN] No tickers to process. Exiting.")
        return

    print("[INFO] Starting EDGAR workflow ...")
    results_df = edgar_workflow.run_workflow(edgar_list, config)
    print(f"[INFO] Workflow completed. Rows: {len(results_df)}")

    # Map back to RIC for reporting, keep EDGAR ticker in a separate column
    try:
        if not results_df.empty and "Ticker" in results_df.columns:
            results_df.insert(0, "EDGAR_Ticker", results_df["Ticker"])
            results_df["Ticker"] = results_df["EDGAR_Ticker"].map(edgar_to_ric).fillna(results_df["EDGAR_Ticker"])
    except Exception as e:
        print(f"[WARN] Unable to map EDGAR tickers back to RICs: {e}")

    # Debug: Inspect columns and a small sample before writing the report
    try:
        debug_mode = config.get("Logging", "level", fallback="INFO").upper() == "DEBUG"
    except Exception:
        debug_mode = False
    if debug_mode:
        try:
            print(f"[DEBUG] Results columns ({len(results_df.columns)}): {list(results_df.columns)}")
            print("[DEBUG] Head sample:")
            print(results_df.head(3).to_string(index=False))
        except Exception as e:
            print(f"[DEBUG] Unable to preview results DataFrame: {e}")

    # Build output path from config
    out_dir_rel = config.get("General", "output_dir", fallback="./output")
    out_dir = (project_root / out_dir_rel).resolve()
    report_name = config.get("General", "report_filename", fallback="AI_Intensity_Report.xlsx")
    out_path = out_dir / report_name

    # Create Excel report
    report_file = reporting.create_excel_report(results_df, out_path)
    print(f"[OK] Report saved to: {report_file}")

    # Verify saved workbook contents (DEBUG only)
    if debug_mode:
        try:
            wb = load_workbook(report_file, read_only=True)
            print(f"[DEBUG] Workbook sheets: {wb.sheetnames}")
            for name in wb.sheetnames:
                ws = wb[name]
                print(f"[DEBUG] Sheet '{name}': rows={ws.max_row}, cols={ws.max_column}, A1='{ws['A1'].value}'")
        except Exception as e:
            print(f"[DEBUG] Unable to inspect saved workbook: {e}")

    # Post-report cleanup: remove the download directory if cleanup is enabled
    cleanup_mode = config.get("EDGAR", "cleanup_filings", fallback="manual").strip().lower()
    if cleanup_mode == "auto":
        dl_dir_rel = config.get("EDGAR", "download_dir", fallback="./sec-edgar-filings")
        dl_dir = (project_root / dl_dir_rel).resolve()
        if dl_dir.exists():
            try:
                remaining = sum(f.stat().st_size for f in dl_dir.rglob("*") if f.is_file())
                shutil.rmtree(dl_dir, ignore_errors=True)
                if remaining > 0:
                    print(f"[OK] Cleaned up download directory ({remaining / 1_048_576:.1f} MB freed)")
                else:
                    print("[OK] Cleaned up empty download directory")
            except Exception as e:
                print(f"[WARN] Could not fully clean download directory: {e}")


if __name__ == "__main__":
    main()
