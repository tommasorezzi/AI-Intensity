"""Excel reporting utilities.

"""

from __future__ import annotations

from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
from datetime import datetime


def _auto_fit_sheet(ws, max_width: int = 60) -> None:
    """Auto-fit column widths, bold header, and enable autofilter on a worksheet."""
    # Bold header row
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Auto filter for the used range
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    # Compute best-fit widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))


def create_excel_report(df: pd.DataFrame, output_path: Union[str, Path]) -> Path:
    """Create a multi-sheet Excel report from the workflow results.

    Sheets:
    - "Riepilogo Generale": grouped sums by Ticker for numeric metrics (excludes Year)
    - "Dati Dettagliati": detailed rows sorted by Ticker and Year
    - "Analisi Trend": pivot of AI_Intensity_Score with Ticker as rows and Year as columns

    Parameters
    ----------
    df : pd.DataFrame
        The detailed results DataFrame (rows per filing). Must include columns
        'Ticker', 'Year', 'AI_Intensity_Score' and keyword count columns.
    output_path : Union[str, Path]
        Destination path for the Excel file. Parent directories will be created
        if they do not exist.

    Returns
    -------
    Path
        The resolved path to the generated Excel file.
    """

    out_path = Path(output_path).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Ensure expected columns exist; proceed gracefully if df is empty or missing
    df = df.copy()

    # Try opening the writer; if target is locked (e.g., open in Excel), fall back to a timestamped file
    target_path = out_path
    try:
        writer = pd.ExcelWriter(target_path, engine="openpyxl")
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_path = out_path.with_name(f"{out_path.stem}_{ts}{out_path.suffix}")
        print(f"[WARN] Output file '{out_path}' is locked. Saving to '{target_path}' instead.")
        writer = pd.ExcelWriter(target_path, engine="openpyxl")

    with writer:
        # Sheet 1: General summary by Ticker (sum numeric columns, excluding Year)
        if not df.empty and "Ticker" in df.columns:
            num_cols = df.select_dtypes(include="number").columns.tolist()
            if "Year" in num_cols:
                num_cols.remove("Year")
            if num_cols:
                general = df.groupby("Ticker", as_index=False)[num_cols].sum()
                # Add filing count so the user knows how many filings contribute
                filing_counts = df.groupby("Ticker").size().reset_index(name="Filing_Count")
                general = general.merge(filing_counts, on="Ticker", how="left")
                # Rename cumulative score to avoid confusion with per-filing score
                if "AI_Intensity_Score" in general.columns:
                    general = general.rename(columns={"AI_Intensity_Score": "AI_Intensity_Score_Total"})
                # Prefer sorting by the total score if available
                sort_by = "AI_Intensity_Score_Total" if "AI_Intensity_Score_Total" in general.columns else num_cols[0]
                general = general.sort_values(sort_by, ascending=False)
            else:
                general = df.drop_duplicates(subset=["Ticker"])[["Ticker"]]
            general.to_excel(writer, index=False, sheet_name="Riepilogo Generale")

        # Sheet 2: Detailed data (sorted)
        detailed = df
        if not detailed.empty:
            # Sort by Ticker and Year when present
            sort_cols = [c for c in ["Ticker", "Year"] if c in detailed.columns]
            if sort_cols:
                detailed = detailed.sort_values(sort_cols)
        detailed.to_excel(writer, index=False, sheet_name="Dati Dettagliati")

        # Sheet 3: Trend analysis pivot (Ticker x Year of AI_Intensity_Score)
        if not df.empty and {"Ticker", "Year", "AI_Intensity_Score"}.issubset(df.columns):
            pivot = df.pivot_table(
                index="Ticker",
                columns="Year",
                values="AI_Intensity_Score",
                aggfunc="sum",
                fill_value=0,
            )
            # Sort index and columns for readability
            pivot = pivot.sort_index()
            try:
                pivot = pivot.reindex(sorted(pivot.columns), axis=1)
            except Exception:
                pass
            pivot.to_excel(writer, sheet_name="Analisi Trend")

    # Re-open the workbook to apply formatting (safer than manipulating during write)
    wb = load_workbook(target_path)
    for ws in wb.worksheets:
        # Bold header and autofilter
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            if ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions

        # Column widths based on cell contents
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx, values_only=True):
                val = row[0]
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

    wb.save(target_path)

    return target_path
